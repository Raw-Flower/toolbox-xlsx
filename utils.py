from django.apps import apps
from django.conf import settings
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tempfile import NamedTemporaryFile
import uuid
from django.core.files.base import File
from pathlib import Path
import os

def getFilePath(instance,filename):
    unique_id = uuid.uuid4().hex
    fileExt = filename[filename.index('.'):]
    newFileName = f'{unique_id}{fileExt}'
    path = f'xlsx/exports/{newFileName}'
    return path

def getImportTemplatePath(instance,filename):
    newFileName = f'import_template.xlsx'
    path = f'import_templates/{instance.app}/{instance.model}/{newFileName}'
    return path

def getCurrentApps():
    current_apps = apps.get_app_configs()
    field_options = [('','-- Select app --')]
    for app in current_apps:
        #if app.path.startswith(str(settings.BASE_DIR)) and app.name != 'xlsx': # Get only the apps located physically in your project
        if app.path.startswith(str(settings.BASE_DIR)): # Get only the apps located physically in your project
            field_options.append((app.name,app.name))
    return sorted(field_options)

def getCurrentModels():
    field_options = []
    for app in apps.get_app_configs():
        if app.path.startswith(str(settings.BASE_DIR)): # Get only the apps located physically in your project
            app_details = apps.get_app_config(app.name)
            for model in app_details.get_models():
                field_options.append((model.__name__,model._meta.verbose_name))
    return field_options

def getModelsByApp(app):
    app_obj = apps.get_app_config(app)
    models_choices = []    
    for model in app_obj.get_models():
        models_choices.append((model.__name__,model._meta.verbose_name))
    return models_choices

def getColumn(pos=0):
    columns = ['A','B','C','D','E','F','G','H','I','J','K','M','N','L','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    return columns[pos]

def getColumnsChoices():
    columns = ['A','B','C','D','E','F','G','H','I','J','K','M','N','L','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    columnsChoices = [(i,i) for i in columns]
    columnsChoices.insert(0,('','--Select column--'))
    return columnsChoices

def sync_def():
    print('Running...')
    time.sleep(2)
    print('Ending...')
    return True

def prepare_xlsx_export(queryset,template_config,log_instance):
    workbook = Workbook() # Create file
    current_sheet = workbook.active # Select sheet
    
    # Header styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="5DE2E7", end_color="5DE2E7", fill_type="solid")
    header_aligment = Alignment(horizontal='center',vertical='center')
    header_border = Border(
        top=Side(border_style='thin',color='000000'),
        right=Side(border_style='thin',color='000000'),
        bottom=Side(border_style='thin',color='000000'),
        left=Side(border_style='thin',color='000000')
    )
   
    # Headers values
    headers_values = [template.label for template in template_config]
    
    # Adding headers
    for col,label in enumerate(headers_values,start=1):
        cell = current_sheet.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_aligment
        cell.border = header_border
    
    # Adding rows
    for record in queryset:
        row = [str(getattr(record,template.value,'')) for template in template_config]
        current_sheet.append(row)
              
    # Saving temp file in log instance
    with NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        workbook.save(temp.name)
        temp.seek(0)
        log_instance.file.save('export.xlsx',File(temp))
        log_instance.status = 2
        log_instance.save()
        
    return log_instance

def create_import_template(config_instance,template_config):
    workbook = Workbook() # Create file
    current_sheet = workbook.active # Select sheet
    
    # Header styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="5DE2E7", end_color="5DE2E7", fill_type="solid")
    header_aligment = Alignment(horizontal='center',vertical='center')
    header_border = Border(
        top=Side(border_style='thin',color='000000'),
        right=Side(border_style='thin',color='000000'),
        bottom=Side(border_style='thin',color='000000'),
        left=Side(border_style='thin',color='000000')
    )
    
    # Headers values
    headers_values = [template.label for template in template_config]
    
    # Adding headers
    for col,label in enumerate(headers_values,start=1):
        cell = current_sheet.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_aligment
        cell.border = header_border
        
    # Saving temp file in configuration instance
    with NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        workbook.save(temp.name)
        temp.seek(0)
        config_instance.import_template.save('import_template.xlsx',File(temp))
        config_instance.template_config = headers_values
        config_instance.save()
    
    #Return configuration instance(should have the template now) 
    return config_instance

def delete_old_import_template(instance):
    file_path = Path(f'{settings.BASE_DIR}/{instance.import_template.url}')
    try:
        os.remove(file_path)
    except FileNotFoundError:
        print(f'ERROR: Image not found in ({file_path}).')  
    except Exception as e:
        print(f'ERROR({type(e).__name__}): {e}')
            
            

    